"""
excel_config.py — Robot Framework Library
==========================================
Reads Environment Variables and Test Scenarios from
  data/TestData.xlsx

Sheets expected:
  • Environments  — rows: Key | DEV | QA | PROD  (column A = key, rest = env values)
  • TestScenarios — rows: TestSuite | TestCase | Username | Password |
                          ExpectedResult | Tags | Active

Usage in Robot Framework:
  Library    ../core/excel_config.ExcelConfig    ENV=DEV
"""

import os
import openpyxl


# Path to the excel file relative to this library's location
_DEFAULT_XLSX = os.path.join(os.path.dirname(__file__), '..', 'data', 'TestData.xlsx')


class ExcelConfig:
    """Robot Framework library for Excel-driven configuration and test data."""

    ROBOT_LIBRARY_SCOPE = 'SUITE'

    def __init__(self, ENV: str = 'DEV', xlsx_path: str = _DEFAULT_XLSX):
        """
        Initialise the Excel config library.

        Args:
            ENV       : Environment name matching a column in the Environments sheet.
                        Valid values: DEV (default), QA, PROD
            xlsx_path : Absolute or relative path to the TestData.xlsx file.
        """
        self._env = ENV.strip().upper()
        self._xlsx_path = os.path.abspath(xlsx_path)
        self._env_vars: dict = {}
        self._test_scenarios: list = []
        self._loaded = False

    # ── Private helpers ──────────────────────────────────────────────────────

    def _load(self):
        """Lazily load data from the Excel workbook."""
        if self._loaded:
            return

        if not os.path.isfile(self._xlsx_path):
            raise FileNotFoundError(
                f"TestData.xlsx not found at: {self._xlsx_path}"
            )

        wb = openpyxl.load_workbook(self._xlsx_path, data_only=True)

        # ── Environments sheet ────────────────────────────────────────────────
        if 'Environments' not in wb.sheetnames:
            raise ValueError("Sheet 'Environments' not found in TestData.xlsx")

        ws_env = wb['Environments']
        header_row = [cell.value for cell in ws_env[1]]

        # Find the column index for our environment (1-indexed)
        try:
            env_col_idx = header_row.index(self._env)  # 0-indexed
        except ValueError:
            available = [h for h in header_row if h and h != 'Key']
            raise ValueError(
                f"Environment '{self._env}' not found in Environments sheet. "
                f"Available: {available}"
            )

        for row in ws_env.iter_rows(min_row=2, values_only=True):
            key = row[0]
            if key is None:
                continue
            value = row[env_col_idx] if env_col_idx < len(row) else None
            self._env_vars[str(key).strip()] = str(value).strip() if value is not None else ''

        # ── TestScenarios sheet ───────────────────────────────────────────────
        if 'TestScenarios' not in wb.sheetnames:
            raise ValueError("Sheet 'TestScenarios' not found in TestData.xlsx")

        ws_ts = wb['TestScenarios']
        ts_headers = [cell.value for cell in ws_ts[1]]
        ts_headers = [str(h).strip() if h else '' for h in ts_headers]

        for row in ws_ts.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            scenario = {ts_headers[i]: (str(v).strip() if v is not None else '')
                        for i, v in enumerate(row) if i < len(ts_headers)}
            self._test_scenarios.append(scenario)

        self._loaded = True

    # ── Public Keywords ──────────────────────────────────────────────────────

    def get_environment_variable(self, key: str) -> str:
        """
        Return the value of an environment variable for the current ENV.

        Example:
            ${url}=    Get Environment Variable    BASE_URL
        """
        self._load()
        if key not in self._env_vars:
            raise KeyError(
                f"Key '{key}' not found in Environments sheet for env '{self._env}'. "
                f"Available keys: {list(self._env_vars.keys())}"
            )
        return self._env_vars[key]

    def get_all_environment_variables(self) -> dict:
        """
        Return ALL environment variables for the current ENV as a dictionary.

        Example:
            ${env}=    Get All Environment Variables
        """
        self._load()
        return dict(self._env_vars)

    def get_test_scenarios(self, suite: str = '', active_only: bool = True) -> list:
        """
        Return test scenario rows from the TestScenarios sheet.

        Args:
            suite       : Filter rows by TestSuite name (case-insensitive). Empty = all.
            active_only : If True (default), only rows where Active == 'YES' are returned.

        Returns:
            List of dicts, each dict representing one row.

        Example:
            ${scenarios}=    Get Test Scenarios    suite=sp_BDD
        """
        self._load()
        result = self._test_scenarios

        if active_only:
            result = [r for r in result if r.get('Active', '').upper() == 'YES']

        if suite:
            result = [r for r in result if r.get('TestSuite', '').lower() == suite.lower()]

        return result

    def load_environment_variables_as_suite_variables(self):
        """
        Load all environment variables from Excel and set them as Robot Framework
        Suite Variables so they are available across all test cases in the suite.

        This keyword is intended for use in ``Suite Setup``.

        Example:
            Suite Setup    Load Environment Variables As Suite Variables
        """
        from robot.libraries.BuiltIn import BuiltIn
        self._load()
        bi = BuiltIn()
        for key, value in self._env_vars.items():
            bi.set_suite_variable(f'${{{key}}}', value)
        bi.log(
            f"Loaded {len(self._env_vars)} environment variables "
            f"from Excel for ENV='{self._env}'",
            level='INFO'
        )

    def log_environment_summary(self):
        """
        Log a summary of all loaded environment variables (masks passwords).

        Example:
            Log Environment Summary
        """
        from robot.libraries.BuiltIn import BuiltIn
        self._load()
        bi = BuiltIn()
        lines = [f"  ENV = {self._env}", f"  XLSX = {self._xlsx_path}", "  Variables:"]
        for key, value in self._env_vars.items():
            masked = '***' if 'PASSWORD' in key.upper() or 'SECRET' in key.upper() else value
            lines.append(f"    {key} = {masked}")
        bi.log('\n'.join(lines), level='INFO', html=False)
