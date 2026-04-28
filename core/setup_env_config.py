import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
EXCEL_PATH = os.path.join(DATA_DIR, 'TestData.xlsx')

def style_header(ws, header_row, bg_color='1F4E79', font_color='FFFFFF'):
    thin = Side(style='thin', color='CCCCCC')
    for cell in ws[header_row]:
        cell.font = Font(name='Calibri', bold=True, color=font_color, size=11)
        cell.fill = PatternFill('solid', fgColor=bg_color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_data_cell(cell, bg='FFFFFF'):
    thin = Side(style='thin', color='CCCCCC')
    cell.font = Font(name='Calibri', size=10)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def setup_excel():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    print(f"Creating Environments sheet...")
    ws_env = wb.create_sheet('Environments')
    ws_env.row_dimensions[1].height = 30

    env_headers = ['Key', 'DEV', 'QA', 'PROD']
    ws_env.append(env_headers)
    style_header(ws_env, 1)

    env_data = [
        ('BASE_URL',         'http://sampleapp.tricentis.com/',              'http://sampleapp.tricentis.com/',              'http://sampleapp.tricentis.com/'),
        ('BROWSER',          'chromium',                                      'chromium',                                      'chromium'),
        ('HEADLESS',         'false',                                         'true',                                          'true'),
        ('API_URL',          'https://restful-booker.herokuapp.com',         'https://restful-booker.herokuapp.com',          'https://restful-booker.herokuapp.com'),
        ('API_USERNAME',     'admin',                                         'admin',                                         'admin'),
        ('API_PASSWORD',     'password123',                                   'password123',                                   'password123'),
        ('LOGIN_URL',        'https://testautomationpractice.blogspot.com/', 'https://testautomationpractice.blogspot.com/', 'https://testautomationpractice.blogspot.com/'),
        ('APP_USERNAME',     'admin',                                         'qa_admin',                                      'prod_admin'),
        ('APP_PASSWORD',     'admin',                                         'qa_pass',                                       'prod_pass'),
        ('TIMEOUT',          '30',                                            '30',                                            '60'),
        ('RETRY_COUNT',      '2',                                             '3',                                             '3'),
        ('SCREENSHOT_ON_FAIL','true',                                         'true',                                          'true'),
    ]

    alt = ['F2F2F2', 'FFFFFF']
    for i, row in enumerate(env_data):
        ws_env.append(list(row))
        bg = alt[i % 2]
        for col in range(1, 5):
            style_data_cell(ws_env.cell(row=i+2, column=col), bg)

    ws_env.column_dimensions['A'].width = 24
    ws_env.column_dimensions['B'].width = 52
    ws_env.column_dimensions['C'].width = 52
    ws_env.column_dimensions['D'].width = 52
    ws_env.freeze_panes = 'B2'

    print(f"Creating TestScenarios sheet...")
    ws_ts = wb.create_sheet('TestScenarios')
    ws_ts.row_dimensions[1].height = 30

    ts_headers = ['TestSuite', 'TestCase', 'Username', 'Password', 'ExpectedResult', 'Tags', 'Active']
    ws_ts.append(ts_headers)
    style_header(ws_ts, 1, bg_color='145A32')

    ts_data = [
        ('sp_BDD',              'Login With Admin',          'admin',           'admin',             'Welcome',  'smoke regression', 'YES'),
        ('sp_BDD',              'Login With Invalid User',   'invalid',         'invalid',           'Error',    'regression',       'YES'),
        ('sp_BDD',              'Login With Empty User',     '',                '',                  'Error',    'edge-case',        'YES'),
        ('Vehicle Insurance',   'Create Quote for Car',      'max.mustermann',  'SecretPassword123!','Success',  'smoke',            'YES'),
        ('Vehicle Insurance',   'Create Quote for Truck',    'max.mustermann',  'SecretPassword123!','Success',  'regression',       'NO'),
        ('Restful Booker',      'Get Bookings',              'admin',           'password123',       '200',      'smoke api',        'YES'),
        ('Restful Booker',      'Create Booking',            'admin',           'password123',       '200',      'smoke api',        'YES'),
        ('Restful Booker',      'Delete Booking',            'admin',           'password123',       '201',      'api',              'YES'),
        ('TestAutomation',      'My First Test',             '',                '',                  'Pass',     'smoke',            'YES'),
        ('TestAutomation',      'Create Random String',      '',                '',                  'Pass',     'regression',       'YES'),
    ]

    alt2 = ['F9FBE7', 'FFFFFF']
    for i, row in enumerate(ts_data):
        ws_ts.append(list(row))
        bg = alt2[i % 2]
        for col in range(1, 8):
            style_data_cell(ws_ts.cell(row=i+2, column=col), bg)

    ws_ts.column_dimensions['A'].width = 22
    ws_ts.column_dimensions['B'].width = 30
    ws_ts.column_dimensions['C'].width = 20
    ws_ts.column_dimensions['D'].width = 22
    ws_ts.column_dimensions['E'].width = 18
    ws_ts.column_dimensions['F'].width = 22
    ws_ts.column_dimensions['G'].width = 10
    ws_ts.freeze_panes = 'A2'

    wb.save(EXCEL_PATH)
    print(f"✅ Setup Complete: {EXCEL_PATH} generated successfully.")

if __name__ == '__main__':
    setup_excel()
