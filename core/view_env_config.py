import os
import sys
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
EXCEL_PATH = os.path.join(DATA_DIR, 'TestData.xlsx')

def view_config():
    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Configuration not found at {EXCEL_PATH}")
        print(f"   Please run `python setup_env_config.py` first.")
        sys.exit(1)

    print(f"Loading Configuration from {EXCEL_PATH}\n")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    if 'Environments' in wb.sheetnames:
        print("====== ENVIRONMENT SETTINGS ======")
        ws = wb['Environments']
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        
        # Determine column width for printing
        col_widths = {i: len(str(h)) for i, h in enumerate(headers)}
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        for row in rows:
            for i, val in enumerate(row):
                if i < len(col_widths):
                    col_widths[i] = max(col_widths[i], len(str(val)) if val is not None else 0)
        
        # Print header
        header_format = " | ".join([f"{{:<{col_widths[i]}}}" for i in range(len(headers))])
        print("-" * (sum(col_widths.values()) + len(headers) * 3))
        print(header_format.format(*headers))
        print("-" * (sum(col_widths.values()) + len(headers) * 3))
        
        # Print rows
        for row in rows:
            formatted = []
            for i in range(len(headers)):
                val = row[i] if i < len(row) and row[i] is not None else ""
                formatted.append(str(val))
            print(header_format.format(*formatted))
            
        print("-" * (sum(col_widths.values()) + len(headers) * 3))
        print("\n")
        
    if 'TestScenarios' in wb.sheetnames:
        print("====== TEST SCENARIOS (ACTIVE ONLY) ======")
        ws = wb['TestScenarios']
        ts_headers = [cell.value for cell in ws[1] if cell.value is not None]
        
        # Display just the key fields: TestSuite, TestCase, ExpectedResult, Active
        active_idx = ts_headers.index('Active') if 'Active' in ts_headers else -1
        
        print(f"{'TestSuite':<22} | {'TestCase':<30} | {'Active'}")
        print("-" * 65)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            active_val = row[active_idx] if active_idx != -1 and active_idx < len(row) else ''
            if str(active_val).strip().upper() == 'YES':
                ts = str(row[0]) if row[0] else ""
                tc = str(row[1]) if row[1] else ""
                print(f"{ts:<22} | {tc:<30} | {active_val}")
        print("-" * 65)

if __name__ == '__main__':
    view_config()
